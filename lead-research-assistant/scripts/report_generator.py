#!/usr/bin/env python3
"""
Lead Report Generator

Creates professional Excel reports from enriched lead data.
Supports multiple templates and visual formatting.
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")


# Color scheme
COLORS = {
    "header": "366092",  # Dark blue
    "tier_a": "00B050",  # Green
    "tier_b": "92D050",  # Light green
    "tier_c": "FFC000",  # Orange
    "tier_d": "FF0000",  # Red
    "alt_row": "F2F2F2",  # Light gray
}


def create_summary_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    """Create executive summary sheet."""
    ws = wb.create_sheet("Executive Summary", 0)

    # Title
    ws["A1"] = "Lead Research Report - Executive Summary"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color=COLORS["header"], fill_type="solid")
    ws.merge_cells("A1:D1")

    # Metadata
    row = 3
    metadata = data.get("metadata", {})
    ws[f"A{row}"] = "Report Date:"
    ws[f"B{row}"] = metadata.get("processed_at", "N/A")
    row += 1
    ws[f"A{row}"] = "Total Leads Analyzed:"
    ws[f"B{row}"] = metadata.get("total_leads", 0)
    row += 1
    ws[f"A{row}"] = "Data Source:"
    ws[f"B{row}"] = metadata.get("input_file", "N/A")

    # Summary stats
    row += 2
    ws[f"A{row}"] = "Performance Metrics"
    ws[f"A{row}"].font = Font(size=14, bold=True)
    row += 1

    summary = data.get("summary", {})

    # Create metrics table
    headers = ["Metric", "Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS["header"], fill_type="solid")

    row += 1
    ws[f"A{row}"] = "Successfully Processed"
    ws[f"B{row}"] = summary.get("successful", 0)
    row += 1
    ws[f"A{row}"] = "Processing Errors"
    ws[f"B{row}"] = summary.get("errors", 0)
    row += 1
    ws[f"A{row}"] = "Average Quality Score"
    ws[f"B{row}"] = summary.get("avg_score", 0)

    # Tier distribution
    row += 2
    ws[f"A{row}"] = "Lead Quality Distribution"
    ws[f"A{row}"].font = Font(size=14, bold=True)
    row += 1

    tier_dist = summary.get("tier_distribution", {})
    headers = ["Tier", "Count", "Percentage"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS["header"], fill_type="solid")

    total_leads = summary.get("successful", 1)
    tiers = [("A", "tier_a"), ("B", "tier_b"), ("C", "tier_c"), ("D", "tier_d")]

    for tier_letter, color_key in tiers:
        row += 1
        count = tier_dist.get(tier_letter, 0)
        percentage = (count / total_leads * 100) if total_leads > 0 else 0

        ws[f"A{row}"] = f"{tier_letter}-Tier"
        ws[f"A{row}"].fill = PatternFill(start_color=COLORS[color_key], fill_type="solid")
        ws[f"B{row}"] = count
        ws[f"C{row}"] = f"{percentage:.1f}%"

    # Top leads
    row += 2
    ws[f"A{row}"] = "Top 10 Priority Leads"
    ws[f"A{row}"].font = Font(size=14, bold=True)
    row += 1

    headers = ["Rank", "Company", "Tier", "Score"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS["header"], fill_type="solid")

    top_leads = summary.get("top_leads", [])
    for idx, lead in enumerate(top_leads, 1):
        row += 1
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = lead.get("company_name", "Unknown")
        ws[f"C{row}"] = lead.get("tier", "D")
        ws[f"D{row}"] = lead.get("score", 0)

        # Color code by tier
        tier = lead.get("tier", "D")
        color_key = f"tier_{tier.lower()}"
        ws[f"C{row}"].fill = PatternFill(start_color=COLORS[color_key], fill_type="solid")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_all_leads_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    """Create detailed leads sheet."""
    ws = wb.create_sheet("All Leads")

    # Headers
    headers = [
        "Company",
        "Tier",
        "Score",
        "Industry",
        "Website",
        "LinkedIn",
        "Contact Name",
        "Contact Title",
        "Status",
        "Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS["header"], fill_type="solid")

    # Data rows
    leads = data.get("leads", [])
    for row_idx, lead in enumerate(leads, 2):
        qual = lead.get("qualification", {})
        tier = qual.get("tier", "D")
        score = qual.get("weighted_total", 0)

        ws[f"A{row_idx}"] = lead.get("company_name", "")
        ws[f"B{row_idx}"] = tier
        ws[f"C{row_idx}"] = round(score, 1)
        ws[f"D{row_idx}"] = lead.get("industry", "")
        ws[f"E{row_idx}"] = lead.get("website", "")
        ws[f"F{row_idx}"] = lead.get("linkedin_url", "")
        ws[f"G{row_idx}"] = lead.get("contact_name", "")
        ws[f"H{row_idx}"] = lead.get("contact_title", "")
        ws[f"I{row_idx}"] = lead.get("status", "")
        ws[f"J{row_idx}"] = lead.get("notes", "")

        # Color code tier column
        color_key = f"tier_{tier.lower()}"
        ws[f"B{row_idx}"].fill = PatternFill(
            start_color=COLORS[color_key], fill_type="solid"
        )

        # Alternate row colors
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                if col != 2:  # Skip tier column
                    ws.cell(row_idx, col).fill = PatternFill(
                        start_color=COLORS["alt_row"], fill_type="solid"
                    )

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def create_qualification_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    """Create detailed qualification scores sheet."""
    ws = wb.create_sheet("Qualification Details")

    # Headers
    headers = [
        "Company",
        "Tier",
        "Total Score",
        "Firmographic",
        "Technographic",
        "Behavioral",
        "Strategic",
        "Recommendation",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS["header"], fill_type="solid")

    # Data rows
    leads = data.get("leads", [])
    for row_idx, lead in enumerate(leads, 2):
        if lead.get("status") != "success":
            continue

        qual = lead.get("qualification", {})

        ws[f"A{row_idx}"] = lead.get("company_name", "")
        ws[f"B{row_idx}"] = qual.get("tier", "D")
        ws[f"C{row_idx}"] = round(qual.get("weighted_total", 0), 1)
        ws[f"D{row_idx}"] = round(
            qual.get("firmographic", {}).get("weighted", 0), 1
        )
        ws[f"E{row_idx}"] = round(
            qual.get("technographic", {}).get("weighted", 0), 1
        )
        ws[f"F{row_idx}"] = round(
            qual.get("behavioral", {}).get("weighted", 0), 1
        )
        ws[f"G{row_idx}"] = round(
            qual.get("strategic", {}).get("weighted", 0), 1
        )
        ws[f"H{row_idx}"] = qual.get("recommendation", "")

        # Color code tier
        tier = qual.get("tier", "D")
        color_key = f"tier_{tier.lower()}"
        ws[f"B{row_idx}"].fill = PatternFill(
            start_color=COLORS[color_key], fill_type="solid"
        )

    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["H"].width = 60
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15

    # Freeze header row
    ws.freeze_panes = "A2"


def create_outreach_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    """Create outreach recommendations sheet."""
    ws = wb.create_sheet("Outreach Strategy")

    # Headers
    headers = [
        "Company",
        "Tier",
        "Priority",
        "Recommended Channel",
        "Messaging Angle",
        "Personalization Hook",
        "Next Steps",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS["header"], fill_type="solid")

    # Generate outreach recommendations based on tier and score
    leads = data.get("leads", [])
    successful_leads = [l for l in leads if l.get("status") == "success"]

    # Sort by score
    sorted_leads = sorted(
        successful_leads,
        key=lambda x: x.get("qualification", {}).get("weighted_total", 0),
        reverse=True,
    )

    for row_idx, lead in enumerate(sorted_leads, 2):
        qual = lead.get("qualification", {})
        tier = qual.get("tier", "D")
        score = qual.get("weighted_total", 0)

        # Determine recommendations based on tier
        if tier == "A":
            priority = "High - Pursue immediately"
            channel = "Multi-touch: LinkedIn + Email + Phone"
            angle = "Executive insight / Strategic partnership"
            hook = "Recent company milestone or funding"
        elif tier == "B":
            priority = "Medium - Engage this quarter"
            channel = "LinkedIn + Email sequence"
            angle = "Problem-solution / ROI focus"
            hook = "Industry trends or competitive intel"
        elif tier == "C":
            priority = "Low - Nurture campaign"
            channel = "Email nurture sequence"
            angle = "Educational content / Thought leadership"
            hook = "Relevant content or case study"
        else:
            priority = "Monitor - Wait for trigger"
            channel = "Passive monitoring"
            angle = "Generic outreach"
            hook = "Funding, hiring, or product launch"

        next_steps = (
            "Research decision-makers → Personalize message → Multi-touch outreach"
            if tier in ["A", "B"]
            else "Add to nurture list → Monitor for signals"
        )

        ws[f"A{row_idx}"] = lead.get("company_name", "")
        ws[f"B{row_idx}"] = tier
        ws[f"C{row_idx}"] = priority
        ws[f"D{row_idx}"] = channel
        ws[f"E{row_idx}"] = angle
        ws[f"F{row_idx}"] = hook
        ws[f"G{row_idx}"] = next_steps

        # Color code tier
        color_key = f"tier_{tier.lower()}"
        ws[f"B{row_idx}"].fill = PatternFill(
            start_color=COLORS[color_key], fill_type="solid"
        )

    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 45

    # Freeze header row
    ws.freeze_panes = "A2"


def create_insights_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    """Create insights and patterns sheet."""
    ws = wb.create_sheet("Insights & Patterns")

    row = 1
    ws[f"A{row}"] = "Lead Research Insights"
    ws[f"A{row}"].font = Font(size=16, bold=True)
    ws.merge_cells(f"A{row}:D{row}")

    row += 2

    # Analyze patterns
    leads = data.get("leads", [])
    successful = [l for l in leads if l.get("status") == "success"]

    # Industry distribution
    ws[f"A{row}"] = "Industry Distribution"
    ws[f"A{row}"].font = Font(size=12, bold=True)
    row += 1

    industries = {}
    for lead in successful:
        industry = lead.get("industry", "Unknown")
        industries[industry] = industries.get(industry, 0) + 1

    ws[f"A{row}"] = "Industry"
    ws[f"B{row}"] = "Count"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].font = Font(bold=True)
    row += 1

    for industry, count in sorted(industries.items(), key=lambda x: x[1], reverse=True)[
        :10
    ]:
        ws[f"A{row}"] = industry
        ws[f"B{row}"] = count
        row += 1

    # Key recommendations
    row += 1
    ws[f"A{row}"] = "Key Recommendations"
    ws[f"A{row}"].font = Font(size=12, bold=True)
    row += 1

    summary = data.get("summary", {})
    tier_dist = summary.get("tier_distribution", {})
    a_tier_count = tier_dist.get("A", 0)
    b_tier_count = tier_dist.get("B", 0)

    recommendations = [
        f"Focus on {a_tier_count} A-tier leads with immediate personalized outreach",
        f"Develop nurture campaigns for {b_tier_count} B-tier leads with educational content",
        "Monitor C/D-tier leads for trigger events (funding, hiring, product launches)",
        "Refine ICP criteria based on highest-converting lead characteristics",
        "Build case studies from similar companies to support outreach",
    ]

    for rec in recommendations:
        ws[f"A{row}"] = f"• {rec}"
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 80

    # Wrap text for recommendations
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True)


def generate_report(
    data: dict[str, Any], output_path: str, template: str = "detailed"
) -> None:
    """
    Generate Excel report from enriched lead data.

    Args:
        data: Enriched lead data dictionary
        output_path: Output file path
        template: Report template (summary, detailed, executive)
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required. Install with: pip install openpyxl"
        )

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create sheets based on template
    if template in ["summary", "detailed", "executive"]:
        create_summary_sheet(wb, data)

    if template in ["detailed", "executive"]:
        create_all_leads_sheet(wb, data)
        create_qualification_sheet(wb, data)
        create_outreach_sheet(wb, data)
        create_insights_sheet(wb, data)

    # Save workbook
    wb.save(output_path)
    print(f"✅ Excel report saved to {output_path}")


def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(
        description="Generate Excel reports from enriched lead data"
    )
    parser.add_argument(
        "--input", type=str, required=True, help="Path to enriched leads JSON file"
    )
    parser.add_argument(
        "--output",
        type=str,
        default="lead_report.xlsx",
        help="Output Excel file path (default: lead_report.xlsx)",
    )
    parser.add_argument(
        "--template",
        type=str,
        choices=["summary", "detailed", "executive"],
        default="detailed",
        help="Report template (default: detailed)",
    )
    parser.add_argument(
        "--top-n",
        type=int,
        default=10,
        help="Number of top leads to highlight (default: 10)",
    )

    args = parser.parse_args()

    # Load enriched data
    try:
        with open(args.input) as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: Input file not found: {args.input}")
        return 1
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in input file: {args.input}")
        return 1

    print(f"📊 Generating {args.template} report from {args.input}...")

    # Generate report
    try:
        generate_report(data, args.output, args.template)

        print("\n✅ Report generation complete!")
        print(f"\n📁 Report saved to: {args.output}")
        print("\n📋 Report includes:")

        if args.template in ["summary", "detailed", "executive"]:
            print("   ✓ Executive Summary - Key metrics and top leads")

        if args.template in ["detailed", "executive"]:
            print("   ✓ All Leads - Complete lead list with scores")
            print("   ✓ Qualification Details - Scoring breakdown")
            print("   ✓ Outreach Strategy - Recommended approaches")
            print("   ✓ Insights & Patterns - Analysis and recommendations")

        print("\n💡 Open the Excel file to explore the full report!\n")

    except Exception as e:
        print(f"Error generating report: {e}")
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
